from datetime import datetime
from pathlib import Path

from flask import Flask, jsonify, request
from openpyxl import Workbook, load_workbook

app = Flask(__name__)

DATA_DIR = Path(__file__).resolve().parent / "data"
WORKBOOK_PATH = DATA_DIR / "user_records.xlsx"

SHEET_HEADERS = {
    "signups": ["first_name", "last_name", "email", "area", "offers", "needs", "created_at"],
    "logins": ["email", "login_status", "created_at"],
    "posts": ["user_name", "type", "category", "title", "area", "availability", "created_at"],
    "reports": [
        "reporter_name",
        "reporter_email",
        "member",
        "listing_ref",
        "report_type",
        "description",
        "incident_date",
        "created_at",
    ],
}


def _ensure_workbook() -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)

    if WORKBOOK_PATH.exists():
        wb = load_workbook(WORKBOOK_PATH)
    else:
        wb = Workbook()
        default_sheet = wb.active
        wb.remove(default_sheet)

    for sheet_name, headers in SHEET_HEADERS.items():
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(title=sheet_name)
            ws.append(headers)

    wb.save(WORKBOOK_PATH)
    wb.close()


def _append_row(sheet_name: str, row_values: list[str]) -> None:
    _ensure_workbook()
    wb = load_workbook(WORKBOOK_PATH)
    ws = wb[sheet_name]
    ws.append(row_values)
    wb.save(WORKBOOK_PATH)
    wb.close()


def _json_or_empty() -> dict:
    return request.get_json(silent=True) or {}


def _now_iso() -> str:
    return datetime.now().isoformat(timespec="seconds")


@app.after_request
def add_cors_headers(response):
    response.headers["Access-Control-Allow-Origin"] = "*"
    response.headers["Access-Control-Allow-Headers"] = "Content-Type"
    response.headers["Access-Control-Allow-Methods"] = "POST, OPTIONS"
    return response


@app.route("/api/<path:_unused>", methods=["OPTIONS"])
def preflight(_unused):
    return ("", 204)


@app.route("/", methods=["GET"])
def home():
    return jsonify(
        {
            "ok": True,
            "message": "Excel API is running",
            "endpoints": [
                "/api/save-signup",
                "/api/save-login",
                "/api/save-post",
                "/api/save-report",
            ],
        }
    )


@app.route("/api/save-signup", methods=["POST"])
def save_signup():
    data = _json_or_empty()
    required = ["first_name", "last_name", "email", "area"]
    if not all(str(data.get(k, "")).strip() for k in required):
        return jsonify({"ok": False, "error": "Missing required signup fields"}), 400

    _append_row(
        "signups",
        [
            str(data.get("first_name", "")).strip(),
            str(data.get("last_name", "")).strip(),
            str(data.get("email", "")).strip(),
            str(data.get("area", "")).strip(),
            str(data.get("offers", "")).strip(),
            str(data.get("needs", "")).strip(),
            _now_iso(),
        ],
    )
    return jsonify({"ok": True})


@app.route("/api/save-login", methods=["POST"])
def save_login():
    data = _json_or_empty()
    email = str(data.get("email", "")).strip()
    if not email:
        return jsonify({"ok": False, "error": "Email is required"}), 400

    _append_row(
        "logins",
        [
            email,
            str(data.get("login_status", "success")).strip() or "success",
            _now_iso(),
        ],
    )
    return jsonify({"ok": True})


@app.route("/api/save-post", methods=["POST"])
def save_post():
    data = _json_or_empty()
    required = ["user_name", "type", "category", "title", "description"]
    if not all(str(data.get(k, "")).strip() for k in required):
        return jsonify({"ok": False, "error": "Missing required post fields"}), 400

    _append_row(
        "posts",
        [
            str(data.get("user_name", "")).strip(),
            str(data.get("type", "")).strip(),
            str(data.get("category", "")).strip(),
            str(data.get("title", "")).strip(),
            str(data.get("area", "")).strip(),
            str(data.get("availability", "")).strip(),
            _now_iso(),
        ],
    )
    return jsonify({"ok": True})


@app.route("/api/save-report", methods=["POST"])
def save_report():
    data = _json_or_empty()
    required = ["reporter_name", "reporter_email", "report_type", "description"]
    if not all(str(data.get(k, "")).strip() for k in required):
        return jsonify({"ok": False, "error": "Missing required report fields"}), 400

    _append_row(
        "reports",
        [
            str(data.get("reporter_name", "")).strip(),
            str(data.get("reporter_email", "")).strip(),
            str(data.get("member", "")).strip(),
            str(data.get("listing_ref", "")).strip(),
            str(data.get("report_type", "")).strip(),
            str(data.get("description", "")).strip(),
            str(data.get("incident_date", "")).strip(),
            _now_iso(),
        ],
    )
    return jsonify({"ok": True})


if __name__ == "__main__":
    _ensure_workbook()
    app.run(debug=True, host="127.0.0.1", port=5000)
